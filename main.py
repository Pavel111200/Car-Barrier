from PIL.Image import ImageTransformHandler
import cv2
import numpy as np
import pytesseract
from ANPR import ANPR
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from license_plate_formats import formats
pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

cascade = cv2.CascadeClassifier("haarcascade_russian_plate_number.xml")
anpr = ANPR()

def extract_license_plate(img_filepath: str) -> dict:
    """
        Extracts the license plate/s of car/s in the given image and returns all the recognized plates in a dictionary

        Parameters:
            img_filename (str):
                image path in string format
    """
    # img=cv2.imread(img_filepath)
    video = cv2.VideoCapture(img_filepath)
    plates = dict()
    while True:
        ret, frame = video.read()

        if ret:
            original_img = frame.copy()
            original_height, original_width, _ = original_img.shape
            frame = cv2.resize(frame, (640, 480))
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            zone = gray[420:480, 0:640]

            nplate = cascade.detectMultiScale(zone, 1.1, 4)
            for (x, y, w, h) in nplate:
                y_increase = original_height / 480
                x_increase = original_width / 640
                plate = original_img[1080 - int((60 - y) * y_increase):1080 - int(((60 - (y + h)) * y_increase)),
                                    int(x * x_increase):int((x + w) * x_increase)]
                
                alpha = 1.3
                beta = 40

                img_thresh = cv2.convertScaleAbs(plate, alpha=alpha, beta=beta)
                cv2.imshow("test", img_thresh)
                cv2.waitKey(0)
                text = anpr.read_text_from_image_easyocr(img_thresh)

                if text is not None:
                    # text = text.replace(" ", "").replace("_", "").upper()
                    if is_valid_license_plate_format(text) and plates.get(text) is None:
                        plates[text] = datetime.now().strftime('%H:%M %d-%m-%Y')
        else:
            return plates

    # gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # nplate = cascade.detectMultiScale(gray, 1.1, 4)

    # for (x, y, w, h) in nplate:
    #     plate = img[y:y+h,x:x+w]
    #     img_thresh = cv2.convertScaleAbs(plate, alpha=2.0, beta=0)
    #     text = anpr.read_text_from_image(img_thresh)

    #     if text is not None and is_valid_license_plate_format(text):
    #         print(text)

def is_valid_license_plate_format(plate_text: str):
    """
        Checks if the given string corresponds to one of the available license plate formats.

        Parameters:
            plate_text (str): the string that needs to be checked  
    """
    for country in formats:
        if re.fullmatch(formats[country], plate_text):
            return True

    return False

def save_dict_to_excel(filepath: str, data: dict):
    """
        Converts the data to a pandas DataFrame and saves it to the given file. If the file does not exist it  will be created

        Parameters:
            filepath (str): path to where the file is located or to be created

            data (dict): the dict object
    """
    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(filepath)

        df = pd.DataFrame(data=data.values(), index=data.keys(),columns=["Time"])
        df.index.name = "Plate"

        with pd.ExcelWriter(filepath, datetime_format="HH:MM DD-MM-YYYY") as writer:
            df.to_excel(writer)
    else:
        sheet = wb.worksheets[0]
        start_row = sheet.max_row

        df = pd.DataFrame(data=data.values(), index=data.keys())

        with pd.ExcelWriter(filepath, datetime_format="HH:MM DD-MM-YYYY", mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, startrow=start_row, header=False)

plates = extract_license_plate("videos/vecteezy_traffic-cars-passing-in-road-with-asphalt-with-cracks-seen_36990287.mov")

save_dict_to_excel("car_entries.xlsx", plates)

# extract_num("images/01576629068.jpeg")
